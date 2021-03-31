# -*- coding: utf-8 -*-
"""
Created on Tue Mar 16 17:33:12 2021

@author: YU Yixiong
"""


import openpyxl
summerWB=openpyxl.load_workbook("C:/Users/YU Yixiong/Desktop/sync/crouseSummer/summer-NameOnly.xlsx")
summerSheet=summerWB["学生课表"]

# for i in range(3,summerSheet.max_row):
#     for j in range(4,summerSheet.max_column):
#         summerSheet.cell(i,j).value=' '
# =============================================================================
# 以上是初始化，避免nonetype的字符串拼接
# =============================================================================
for w in range(5):
    path=" ("+str(w+1)+").xlsx"
    wb=openpyxl.load_workbook(path)
    Sheet=wb["学生课表"]
            # summerSheet.cell(i,j).value+=Sheet.cell(i,j).value
            # print(Sheet.cell(i,j).value)
        
    for i in range(3,11):
        for j in range(4,9):
            # summerSheet.cell(i,j).value+="\n"
            if Sheet.cell(i,j).value!=None:
                Sheet.cell(i,j).value=Sheet.cell(i,j).value.replace("\n", "")
                name=''
                for t in range(len(Sheet.cell(i,j).value)):
                    if Sheet.cell(i,j).value[t]==' ':
                        name=Sheet.cell(i,j).value[0:t]
                        break
                summerSheet.cell(i,j).value+=name+" "
                print(summerSheet.cell(i,j).value)
    
    wb.close()

summerWB.save("summer-NameOnly.xlsx")