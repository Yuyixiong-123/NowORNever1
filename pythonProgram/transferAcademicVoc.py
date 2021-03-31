# -*- coding: utf-8 -*-
"""
Created on Tue Mar 30 20:23:00 2021

@author: YU Yixiong
"""

import openpyxl
import http.client
import hashlib
import urllib
import random
import json

def transfer(q):
     appid = '20210319000733308'        # 填写你的appid
     secretKey = 'Uj1jSmjNJhm6ImkcZyFC'    # 填写你的密钥
     
     httpClient = None
     myurl = '/api/trans/vip/translate'  # 通用翻译API HTTP地址
      
     fromLang = 'auto'       # 原文语种
     toLang = 'zh'           # 译文语种
     salt = random.randint(32768, 65536)
     sign = appid + q + str(salt) + secretKey
     sign = hashlib.md5(sign.encode()).hexdigest()
     myurl = myurl + '?appid=' + appid + '&q=' + urllib.parse.quote(q) + '&from=' + fromLang + \
             '&to=' + toLang + '&salt=' + str(salt) + '&sign=' + sign
     # f=open("transferResult.txt","a+")
     # f.write('\n\n'+q+'\n')
     try:
         httpClient = http.client.HTTPConnection('api.fanyi.baidu.com')
         httpClient.request('GET', myurl)
         # response是HTTPResponse对象
         response = httpClient.getresponse()
         result_all = response.read().decode("utf-8")
         result = json.loads(result_all)
         
         # return(result)
      
         # f.write (result["trans_result"][0]["dst"])
         # f.close()
         return result["trans_result"][0]["dst"]
      
     except Exception as e:
         print (e)
     finally:
         if httpClient:
             httpClient.close()

wb=openpyxl.load_workbook("list1-3.xlsx")
sheet=wb["Sheet1"]
for col in (1,3,5):
    for i in range(1,sheet.max_row):
        sheet.cell(i,col+1).value=transfer(sheet.cell(i,col).value)

wb.save("list1-3.xlsx")