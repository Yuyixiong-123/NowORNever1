# -*- coding: utf-8 -*-
"""
Created on Sat Feb 27 11:04:35 2021

@author: yixiong YU

note: 
"""
xuehaoCol='C'
path="C:/Users/YuYX/Desktop/附件2：疫苗接种时间安排.xlsx"
sheetname='Sheet1'

import openpyxl

wb=openpyxl.load_workbook(path)
sheet=wb[sheetname]

for i in range(2,sheet.max_row):
     if(sheet.cell(i,3).value[:6]!='BY2005' and sheet.cell(i,3).value[:6]!='ZB2005'):
          sheet.cell(i,1).value=None

wb.save(path)